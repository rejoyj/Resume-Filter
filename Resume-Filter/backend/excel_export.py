import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class ExcelExporter:
    def __init__(self):
        self.header_style = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        self.cell_style = {
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def export_to_excel(self, parsed_resumes, output_file):
        """Export parsed resume data to Excel with formatting"""
        try:
            # Prepare data for DataFrame
            df_data = []
            
            for resume in parsed_resumes:
                row = {
                    'File Name': resume.get('file_name', ''),
                    'Name': resume.get('name', ''),
                    'Email': resume.get('email', ''),
                    'Phone Number': resume.get('phone_number', ''),
                    'Skills': self._format_list_field(resume.get('skills')),
                    'Education': self._format_list_field(resume.get('education')),
                    'Location': resume.get('location', ''),
                    'Total Experience (Years)': resume.get('total_experience', ''),
                    'Processing Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                df_data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(df_data)
            
            # Create Excel writer
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Write main data
                df.to_excel(writer, sheet_name='Resume Data', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Resume Data']
                
                # Apply formatting
                self._format_worksheet(worksheet, len(df))
                
                # Create summary sheet
                self._create_summary_sheet(workbook, parsed_resumes)
                
                # Create skills analysis sheet
                self._create_skills_analysis_sheet(workbook, parsed_resumes)
            
            print(f"✓ Excel file exported successfully: {output_file}")
            return True
            
        except Exception as e:
            print(f"❌ Error exporting to Excel: {str(e)}")
            return False

    def _format_list_field(self, field_data):
        """Format list fields for Excel display"""
        if field_data and isinstance(field_data, list):
            return ', '.join(str(item) for item in field_data)
        elif field_data:
            return str(field_data)
        return ''

    def _format_worksheet(self, worksheet, data_rows):
        """Apply formatting to the main worksheet"""
        # Set column widths
        column_widths = {
            'A': 25,  # File Name
            'B': 20,  # Name
            'C': 30,  # Email
            'D': 15,  # Phone
            'E': 50,  # Skills
            'F': 40,  # Education
            'G': 25,  # Location
            'H': 15,  # Experience
            'I': 20   # Processing Date
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
            cell.border = self.header_style['border']
        
        # Apply cell formatting
        for row in range(2, data_rows + 2):
            for col in range(1, 10):  # A to I
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = self.cell_style['alignment']
                cell.border = self.cell_style['border']
        
        # Set row heights for better readability
        for row in range(2, data_rows + 2):
            worksheet.row_dimensions[row].height = 60

    def _create_summary_sheet(self, workbook, parsed_resumes):
        """Create a summary statistics sheet"""
        summary_sheet = workbook.create_sheet('Summary')
        
        # Calculate statistics
        total_resumes = len(parsed_resumes)
        stats = {
            'Total Resumes Processed': total_resumes,
            'Resumes with Name': sum(1 for r in parsed_resumes if r.get('name')),
            'Resumes with Email': sum(1 for r in parsed_resumes if r.get('email')),
            'Resumes with Phone': sum(1 for r in parsed_resumes if r.get('phone_number')),
            'Resumes with Skills': sum(1 for r in parsed_resumes if r.get('skills')),
            'Resumes with Education': sum(1 for r in parsed_resumes if r.get('education')),
            'Resumes with Location': sum(1 for r in parsed_resumes if r.get('location')),
            'Resumes with Experience': sum(1 for r in parsed_resumes if r.get('total_experience'))
        }
        
        # Add headers
        summary_sheet['A1'] = 'Statistic'
        summary_sheet['B1'] = 'Count'
        summary_sheet['C1'] = 'Percentage'
        
        # Add data
        row = 2
        for stat_name, count in stats.items():
            summary_sheet[f'A{row}'] = stat_name
            summary_sheet[f'B{row}'] = count
            if stat_name != 'Total Resumes Processed':
                percentage = (count / total_resumes * 100) if total_resumes > 0 else 0
                summary_sheet[f'C{row}'] = f"{percentage:.1f}%"
            else:
                summary_sheet[f'C{row}'] = "100%"
            row += 1
        
        # Format summary sheet
        for cell in summary_sheet[1]:
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
        
        # Set column widths
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 15
        summary_sheet.column_dimensions['C'].width = 15

    def _create_skills_analysis_sheet(self, workbook, parsed_resumes):
        """Create a skills analysis sheet"""
        skills_sheet = workbook.create_sheet('Skills Analysis')
        
        # Collect all skills
        skills_count = {}
        for resume in parsed_resumes:
            skills = resume.get('skills', [])
            if skills:
                for skill in skills:
                    skill_lower = skill.lower()
                    skills_count[skill_lower] = skills_count.get(skill_lower, 0) + 1
        
        # Sort skills by frequency
        sorted_skills = sorted(skills_count.items(), key=lambda x: x[1], reverse=True)
        
        # Add headers
        skills_sheet['A1'] = 'Skill'
        skills_sheet['B1'] = 'Frequency'
        skills_sheet['C1'] = 'Percentage'
        
        # Add top skills (limit to top 50)
        total_resumes = len(parsed_resumes)
        for i, (skill, count) in enumerate(sorted_skills[:50], 2):
            skills_sheet[f'A{i}'] = skill.title()
            skills_sheet[f'B{i}'] = count
            percentage = (count / total_resumes * 100) if total_resumes > 0 else 0
            skills_sheet[f'C{i}'] = f"{percentage:.1f}%"
        
        # Format skills sheet
        for cell in skills_sheet[1]:
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
        
        # Set column widths
        skills_sheet.column_dimensions['A'].width = 25
        skills_sheet.column_dimensions['B'].width = 15
        skills_sheet.column_dimensions['C'].width = 15

def export_to_excel_with_formatting(parsed_resumes, output_file):
    """Standalone function for enhanced Excel export"""
    exporter = ExcelExporter()
    return exporter.export_to_excel(parsed_resumes, output_file)